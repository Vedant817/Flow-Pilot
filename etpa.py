import time
import openpyxl
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Define the file path and the directory to watch
excel_file_path = r'D:\Deloitte\Prototype\Sample.xlsx'  # Replace with your file path
directory_to_watch = r'D:\Deloitte\Prototype'  # Replace with the directory of the file

# Store previous content for comparison
previous_content = {}

# Function to read the Excel file and return content as a dictionary
def read_excel_file():
    content = {}
    try:
        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active
        
        # Read the content of the entire sheet into a dictionary
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                content[cell.coordinate] = cell.value
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
    return content

# Function to compare previous content with the new content
def compare_changes(new_content):
    changes = {}
    
    # Check for changes in the content
    for cell, new_value in new_content.items():
        if cell not in previous_content or previous_content[cell] != new_value:
            changes[cell] = new_value
    
    # Update previous content with new content for future comparisons
    previous_content.update(new_content)
    
    return changes

# Function to handle modified part of the file
def handle_modified_file():
    new_content = read_excel_file()
    changes = compare_changes(new_content)
    
    if changes:
        print(f"Detected changes in the file:")
        for cell, new_value in changes.items():
            print(f"Cell {cell} changed to {new_value}")
    else:
        print("No changes detected.")

# Create an event handler that responds to file modifications
class ExcelFileHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path == excel_file_path:
            print(f"Detected change in file: {excel_file_path}")
            handle_modified_file()

# Set up the observer to watch for file modifications
event_handler = ExcelFileHandler()
observer = Observer()
observer.schedule(event_handler, directory_to_watch, recursive=False)

# Start the observer
observer.start()
print(f"Monitoring changes in {excel_file_path}...")

# Run the monitoring loop
try:
    while True:
        time.sleep(1)  # Keeps the script running and monitoring for changes
except KeyboardInterrupt:
    observer.stop()
    print("Stopped monitoring.")

# Stop the observer when done
observer.join()
